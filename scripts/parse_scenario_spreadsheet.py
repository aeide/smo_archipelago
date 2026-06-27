#!/usr/bin/env python3
"""
parse_scenario_spreadsheet.py

Compile Devon's authored ground-truth spreadsheet
(``Odyssey Scenario_Gating Logic.xlsx``, repo root) into a committed,
IP-safe scenario-gate table:

    apworld/smo_archipelago/data/scenario_gates.json
        { "<Kingdom>: <Moon>": "<requires fragment>", ... }

WHY this exists (read before touching the scenario layer):
  The romfs ``progress_bit_flag`` (used by compile_moon_logic.py's bit-driven
  tiers) measures a moon's OBJECT PRESENCE across scenario layouts, not its
  COLLECTABILITY.  Dozens of moons whose object exists in the arrival layout
  (bit-0 set) are not actually reachable until a story event fires — so the
  bit-driven pass shipped them FREE while the spreadsheet correctly gates them
  "Available after collecting <story moon>".  The spreadsheet is the verified
  source of truth; this script makes it the authority for scenario gating.

IP posture (matches compile_moon_logic.py): the spreadsheet's natural-language
availability TEXT never lands in the output.  Only functional predicates
(``{<Kingdom>Peace()}`` / ``{canReachLocation(<committed location name>)}``) and
committed location names are emitted — the exact same shapes already living in
locations.json ``requires``.  Safe to commit.

CARVE-OUTS (deliberately NOT emitted here — compile_moon_logic.py keeps its
existing handling for these):
  * Cascade — its clear scenario is its LAST, so it gates on the dedicated
    {CascadeDeparture()} / {CascadePeace()} pass (leave-deadlock fix), which the
    flat "after <Multi Moon>" sheet text cannot express.  The one Cascade row we
    DO emit is the fork painting "Secret Path to Fossil Falls!" (order-independent
    gate), which is independent of the departure logic.
  * Moon Kingdom — its "after defeating Bowser in this kingdom" layer is the
    runtime moon_postwin filler restriction, gated on the game-clear goal.

The emitted fragment is ANDed ONTO each moon's move-set / capture requirement by
compile_moon_logic.py (it does not replace it).

Run on the machine that has the spreadsheet (Devon's), then commit the JSON:
    python scripts/parse_scenario_spreadsheet.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
SPREADSHEET = REPO_ROOT / "Odyssey Scenario_Gating Logic.xlsx"
DATA_DIR = REPO_ROOT / "apworld" / "smo_archipelago" / "data"
LOCATIONS = DATA_DIR / "locations.json"
OUT = DATA_DIR / "scenario_gates.json"

# Sheet name -> location-name prefix (the part before ": " in locations.json).
SHEET_PREFIX = {
    "Cap": "Cap", "Cascade": "Cascade", "Sand": "Sand", "Lake": "Lake",
    "Wooded": "Wooded", "Cloud": "Cloud", "Lost": "Lost", "Metro": "Metro",
    "Snow": "Snow", "Seaside": "Seaside", "Luncheon": "Luncheon",
    "Ruined": "Ruined", "Bowsers": "Bowser's", "Moon": "Moon",
    "Mushroom": "Mushroom", "Dark Side": "Dark Side",
}

# {<Kingdom>Peace()} predicate per location prefix.  Bowser's uses BowserPeace;
# Dark Side / Mushroom have no boss-peace predicate (their gates are all
# canReachLocation()-based).
PEACE = {
    "Cap": "{CapPeace()}", "Cascade": "{CascadePeace()}", "Sand": "{SandPeace()}",
    "Lake": "{LakePeace()}", "Wooded": "{WoodedPeace()}", "Cloud": "{CloudPeace()}",
    "Lost": "{LostPeace()}", "Metro": "{MetroPeace()}", "Snow": "{SnowPeace()}",
    "Seaside": "{SeasidePeace()}", "Luncheon": "{LuncheonPeace()}",
    "Ruined": "{RuinedPeace()}", "Bowser's": "{BowserPeace()}",
    "Moon": "{MoonPeace()}",
}

# Kingdoms whose scenario gating is owned elsewhere (see module docstring).
CARVE_OUT = {"Cascade", "Moon"}

# Fork-order-dependent painting warps (docs/scenario-gating-audit-todo.md §11).
# Order-independent gate: true only when the moon is reachable in EVERY consistent
# fork order.  Keyed by full location name; applied regardless of CARVE_OUT (the
# Cascade entry here is the painting, independent of the departure pass).
FORK_GATES = {
    "Cascade: Secret Path to Fossil Falls!": "({SnowPeace()} or {SeasidePeace()})",
    "Lake: Secret Path to Lake Lamode!": "({SnowPeace()} or {SeasidePeace()})",
    "Wooded: Secret Path to the Steam Gardens!": "({SnowPeace()} or {SeasidePeace()})",
    "Sand: Secret Path to Tostarena!": "{canReachLocation(Wooded: Flower Thieves of Sky Garden)}",
    "Luncheon: Secret Path to Mount Volbono!": "{canReachLocation(Wooded: Flower Thieves of Sky Garden)}",
    "Snow: Secret Path to Shiveria!": "{MoonPeace()}",
    "Seaside: Secret Path to Bubblaine!": "{MoonPeace()}",
}

KINGDOM_WORDS = ("Cap", "Cascade", "Sand", "Lake", "Wooded", "Cloud", "Lost",
                 "Metro", "Snow", "Seaside", "Luncheon", "Ruined", "Bowser's",
                 "Bowser", "Moon", "Mushroom", "Dark Side")


def clean_name(s: str) -> str:
    """Strip the spreadsheet's ❸ multi-moon marker / leading asterisks / extra
    whitespace from a moon name so it matches the locations.json spelling."""
    s = s.replace("❶", "").replace("❷", "").replace("❸", "")
    s = s.replace("❹", "").replace("❺", "")
    s = s.lstrip("*").strip()
    return re.sub(r"\s+", " ", s)


def _aggr(s: str) -> str:
    """Aggressive normal form: lowercase, strip every non-alphanumeric char.
    Absorbs the locations.json spelling quirks that differ from the sheet —
    a stray leading apostrophe (``'Round-the-World Tourist``), an inconsistent
    trailing ``!``, and Luncheon's extra colon (``Kingdom: Regular Cup``)."""
    return re.sub(r"[^a-z0-9]", "", s.lower())


class Resolver:
    """Punctuation-tolerant moon-name -> committed location-name resolver."""

    def __init__(self, location_names: set[str]):
        self.by_lower = {n.lower(): n for n in location_names}
        self.names = location_names
        # Aggressive-normalized fallback index, but only for forms that are
        # UNIQUE (so we never silently map two distinct moons together).
        seen: dict[str, str | None] = {}
        for n in location_names:
            k = _aggr(n)
            seen[k] = None if k in seen else n
        self.by_aggr = {k: v for k, v in seen.items() if v is not None}

    def resolve(self, prefix: str, raw: str) -> str | None:
        cand = f"{prefix}: {clean_name(raw)}"
        hit = self.by_lower.get(cand.lower())
        if hit:
            return hit
        return self.by_aggr.get(_aggr(cand))

    def kingdom_members(self, prefix: str, contains: str) -> list[str]:
        pl = f"{prefix}: ".lower()
        return sorted(n for n in self.names
                      if n.lower().startswith(pl) and contains.lower() in n.lower())


def canreach(name: str) -> str:
    return f"{{canReachLocation({name})}}"


def and_join(parts: list[str]) -> str:
    parts = [p for p in parts if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return "(" + " and ".join(parts) + ")"


# Match: collecting [the] "X" [Power|Multi] Moon|Multi ... optionally (K Kingdom)
_QUOTED = re.compile(r'collecting (?:the )?"([^"]+)"(?:\s+(?:Power|Multi)\s+\w+)?'
                     r'(?:\s*\(([^)]*?)Kingdom\))?', re.IGNORECASE)


def parse_row(prefix: str, name: str, text: str, R: Resolver,
              warnings: list[str]) -> str | None:
    """Return the scenario requires fragment for one moon, or None for FREE /
    not-applicable (no gate emitted)."""
    full = f"{prefix}: {clean_name(name)}"
    low = text.lower()

    # 1) Fork painting warps — hardcoded order-independent gate.  Applied even
    #    for the carve-out kingdoms (Cascade's Fossil Falls painting is
    #    independent of its departure logic).
    if full in FORK_GATES:
        return FORK_GATES[full]

    # 1b) Cascade / Moon are owned by compile_moon_logic.py's dedicated passes
    #     (CascadeDeparture pass; Moon postwin/rearrival).  Emit NOTHING else for
    #     them so the spreadsheet layer never clobbers that load-bearing handling.
    if prefix in CARVE_OUT:
        return None

    # 2) Skilled-jump soft gates -> intentionally FREE (logic floor, §5).
    if "skilled jump" in low:
        return None

    # 3) Anything available "from the start" (incl. pyramid-toggle moons that
    #    START available then flicker) is FREE for minimal accessibility (§5).
    if low.startswith("available from the start"):
        return None

    # 4) Odyssey power-up / repair => de-facto free (kingdom already reached).
    if "powering up the odyssey" in low or "repairing the odyssey" in low:
        return None

    # 5) Moon Kingdom "defeating Bowser in THIS kingdom" => carve-out (postwin).
    if "in this kingdom" in low:
        return None

    # 5b) "leaving the <K> Kingdom and then returning" => re-arrival layer =>
    #     kingdom peace ({LostPeace()} etc.).  Cascade's three re-arrival rows are
    #     a carve-out (their PEACE() resolves but the dedicated pass owns Cascade).
    if "leaving the" in low and "returning" in low:
        if prefix in CARVE_OUT:
            return None
        return PEACE.get(prefix) or None

    # 6) Moon Rock open => kingdom peace (+ optional "and collecting X" Cup gate).
    if "moon rock" in low:
        frag = PEACE.get(prefix, "")
        m = re.search(r'and collecting (?:the )?"([^"]+)"', text, re.IGNORECASE)
        if m:
            anchor = R.resolve(prefix, m.group(1))
            if anchor:
                return and_join([frag, canreach(anchor)])
            warnings.append(f"{full}: Cup anchor unresolved {m.group(1)!r}")
        return frag or None

    # 7) "defeating Bowser in the Moon Kingdom" (moonclear) => own-kingdom peace
    #    (our moon-rock system surfaces these at the kingdom's own peace).
    if "in the moon kingdom" in low:
        return PEACE.get(prefix) or None

    # 8) Ruined Dragon boss => kingdom peace.
    if "ruined dragon" in low:
        return PEACE.get(prefix) or None

    # 9) Toad outside Peach's Castle (World-Traveling Peach) => kingdom peace.
    if "peach's castle" in low and "secret path" not in full.lower():
        return PEACE.get(prefix) or None

    # 10) "all 4 Power Moons from the band members / barriers" / "four seals" =>
    #     AND canReachLocation(each prerequisite moon).
    if "band members" in low:
        members = R.kingdom_members(prefix, "on Board")
        return and_join([canreach(m) for m in members]) or None
    if "from the barriers" in low:
        members = R.kingdom_members(prefix, "Barrier")
        return and_join([canreach(m) for m in members]) or None
    if "four seals" in low:
        members = R.kingdom_members(prefix, "Seal")
        return and_join([canreach(m) for m in members]) or None

    # 11) Generic "collecting [the] "X" [Power/Multi Moon]" — possibly several
    #     (cross-kingdom rows carry a "(K Kingdom)" qualifier; same-kingdom rows
    #     do not).  AND every resolved anchor (cross-kingdom Tourist rows chain
    #     two; story rows have one).  Cascade carve-out is applied after fork.
    anchors: list[str] = []
    for raw, kq in _QUOTED.findall(text):
        kq = kq.strip()
        target_prefix = prefix
        if kq:
            kq = kq.strip()
            # normalize "Bowser's"/"Dark Side" spellings already in KINGDOM_WORDS
            for kw in KINGDOM_WORDS:
                if kq.lower().startswith(kw.lower()):
                    target_prefix = "Bowser's" if kw == "Bowser" else kw
                    break
        a = R.resolve(target_prefix, raw)
        if a:
            anchors.append(canreach(a))
        else:
            warnings.append(f"{full}: anchor unresolved {raw!r} (kingdom {target_prefix})")
    if anchors:
        if prefix in CARVE_OUT:
            return None  # Cascade story moons owned by the dedicated pass
        # dedupe preserving order
        seen: set[str] = set()
        anchors = [a for a in anchors if not (a in seen or seen.add(a))]
        return and_join(anchors)

    # Unclassified -> leave free but warn so the grammar stays covered.
    if prefix not in CARVE_OUT:
        warnings.append(f"{full}: UNCLASSIFIED text {text!r}")
    return None


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required: pip install openpyxl", file=sys.stderr)
        return 2
    if not SPREADSHEET.exists():
        print(f"spreadsheet not found: {SPREADSHEET}", file=sys.stderr)
        return 2

    locations = json.loads(LOCATIONS.read_text(encoding="utf-8"))
    R = Resolver({l["name"] for l in locations})

    wb = openpyxl.load_workbook(SPREADSHEET, read_only=True, data_only=True)
    gates: dict[str, str] = {}
    warnings: list[str] = []
    seen_moons = 0
    unmatched_moons: list[str] = []

    for sheet in wb.sheetnames:
        prefix = SHEET_PREFIX.get(sheet)
        if prefix is None:
            warnings.append(f"unknown sheet {sheet!r}")
            continue
        for row in wb[sheet].iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            text = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not text:
                continue
            seen_moons += 1
            # Resolve the moon's OWN name to its canonical locations.json spelling
            # (punctuation-tolerant) so the emitted KEY byte-matches the file.
            full = R.resolve(prefix, name)
            if full is None:
                cand = f"{prefix}: {clean_name(name)}"
                if cand in FORK_GATES:
                    full = cand
                else:
                    unmatched_moons.append(cand)
                    continue
            frag = parse_row(prefix, name, text, R, warnings)
            if frag:
                gates[full] = frag

    OUT.write_text(
        json.dumps(dict(sorted(gates.items())), indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8")

    print(f"spreadsheet rows parsed:  {seen_moons}")
    print(f"scenario gates emitted:   {len(gates)} -> {OUT.relative_to(REPO_ROOT)}")
    # breakdown
    npeace = sum(1 for v in gates.values() if "Peace()" in v and "canReach" not in v)
    ncr = sum(1 for v in gates.values() if "canReachLocation" in v)
    print(f"  pure peace:             {npeace}")
    print(f"  canReachLocation-based: {ncr}")
    if unmatched_moons:
        print(f"\n  ⚠ {len(unmatched_moons)} spreadsheet moons not found in locations.json:")
        for m in unmatched_moons:
            print(f"      {m}")
    if warnings:
        print(f"\n  ⚠ {len(warnings)} warnings:")
        for w in warnings:
            print(f"      {w}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
