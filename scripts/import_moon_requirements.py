#!/usr/bin/env python3
"""
import_moon_requirements.py

Parse the corrected "SMO Requirements.xlsx" (sheet "Moons") into two committed
JSON files:

  apworld/smo_archipelago/data/moon_requirements.json
      Keyed by sheet name.  Each entry has a ``location_name`` field giving
      the matched locations.json canonical name (null when unmatched), plus the
      faithful structured requirements (capture groups + per-method jump/throw/
      other terms).  This is the "what the sheet says" layer — the logic
      compiler (Stage 2) turns it into AP ``requires`` strings.

  apworld/smo_archipelago/data/subareas.json
      Keyed by subarea prefix (e.g. "Frog Pond").  Each entry records the
      parent kingdom and the sheet + location names of its moons.

Sheet layout (1-indexed columns):
    A  ignorable (done checkbox)
    B  moon name "<Kingdom or Subarea>: <Moon Name>"  (merged over the 3-row block)
    C  required capture(s)                              (merged when a single
       AND-group; broken into the 3 sub-rows when there are OR alternatives)
    D  row label: "Minimum Jump Height:" / "Cap Throws that work:" / "Other Required:"
    E-I  Method 1..5 (ORed alternatives), each a 3-row (jump / throws / other) cell
    J  ignorable (human counters)

Capture encoding (load-bearing):
    Commas WITHIN one C cell  -> AND (all required together)
    Separate C sub-cells       -> OR  (alternative ways to satisfy the capture)
    A literal "None" C sub-cell -> a no-capture alternative (capture_optional)
    "Locked behind Default Capture" -> stage's own default capture (sentinel)

Usage (from repo root, Windows — needs openpyxl):
    python scripts/import_moon_requirements.py
    python scripts/import_moon_requirements.py path/to/custom.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "apworld" / "smo_archipelago" / "data"
DEFAULT_XLSX = REPO_ROOT / "SMO Requirements.xlsx"
SHEET_NAME = "Moons"
OUT_REQUIREMENTS = DATA_DIR / "moon_requirements.json"
OUT_SUBAREAS = DATA_DIR / "subareas.json"
ENTRANCE_STAGES = DATA_DIR / "entrance_stages.json"

# ─────────────────────────────────────────────────────────────────────────────
# Kingdom prefix normalisation
# Maps the long sheet kingdom prefix to the abbreviated locations.json prefix.
# These are the ONLY prefixes treated as "main kingdom" entries; everything
# else is a subarea.
# ─────────────────────────────────────────────────────────────────────────────
KINGDOM_PREFIX_MAP: dict[str, str] = {
    "Cap Kingdom":      "Cap",
    "Cascade Kingdom":  "Cascade",
    "Sand Kingdom":     "Sand",
    "Wooded Kingdom":   "Wooded",
    "Lake Kingdom":     "Lake",
    "Cloud Kingdom":    "Cloud",
    "Lost Kingdom":     "Lost",
    "Metro Kingdom":    "Metro",
    "Snow Kingdom":     "Snow",
    "Seaside Kingdom":  "Seaside",
    "Luncheon Kingdom": "Luncheon",
    "Ruined Kingdom":   "Ruined",
    "Bowser's Kingdom": "Bowser's",
    "Moon Kingdom":     "Moon",
    "Mushroom Kingdom": "Mushroom",
    "Dark Side":        "Dark Side",
    "Darker Side":      "Darker Side",
}

MAIN_KINGDOM_PREFIXES: frozenset[str] = frozenset(KINGDOM_PREFIX_MAP)

# ─────────────────────────────────────────────────────────────────────────────
# Vocabulary maps
# ─────────────────────────────────────────────────────────────────────────────

# Minimum jump height  (row 0, cols E-I).  Enum carries the in-game height so the
# Stage-2 compiler can build the "OR of every jump item reaching >= X" ladder.
JUMP_HEIGHT_MAP: dict[str, str] = {
    "No Jump Needed":                 "none",
    "Single Jump (258)":              "single",
    "Double Jump (312)":              "double",
    "Cap Return Jump (400)":          "cap_return",
    "Backflip/Vault/Side Flip (496)": "backflip",   # Backflip OR Side Flip OR Cap Bounce(vault)
    "Ground Pound Jump (514)":        "gpj",
    "Triple Jump (550)":              "triple",
    "Long Jump":                      "long_jump",  # own horizontal axis
}

# Cap throws (row 1, cols E-I — comma-separated within each cell)
CAP_THROW_MAP: dict[str, str] = {
    "No Cap Throw Needed": "none",
    "Neutral Throw":       "neutral",
    "Up Throw":            "up",
    "Down Throw":          "down",
    "Spin Throw":          "spin",
}

# Other requirements (row 2, cols E-I — comma-separated within each cell).
# "None" → skipped (empty); "Capture" → requires one of the col-C capture groups.
OTHER_REQUIRED_MAP: dict[str, str | None] = {
    "None":          None,
    "Capture":       "capture",
    "Ground Pound":  "ground_pound",
    "Dive":          "dive",
    "Wall Slide":    "wall_slide",   # new in corrected sheet
    "Climb":         "climb",
    "Roll":          "roll",
    "Crouch":        "crouch",
    "Roll Boost":    "roll_boost",
    "Cap Bounce":    "cap_bounce",   # new in corrected sheet
    "Bonk (Roll)":   "bonk_roll",
    # A stray jump-height token occasionally appears under Other Required.
    "Single Jump (258)": "single",
}

# Sentinel value for col C when the stage uses its default capture
LOCKED_SENTINEL = "Locked behind Default Capture"

# Sheet capture display-name → AP items.json capture name.  Only the handful that
# disagree need entries; everything else passes through unchanged.
CAPTURE_NAME_MAP: dict[str, str] = {
    "Wiggler":     "Tropical Wiggler",
    "Ty-Foo":      "Ty-foo",
    "Spark Pylon": "Spark pylon",   # one cell capitalises; canonical is lowercase 'pylon'
}

# One-off sheet name → locations.json canonical name overrides.
CSV_NAME_OVERRIDES: dict[str, str] = {
    # Sheet uses a sub-qualifier "Upper Interior:" not present in locations.json
    "Inverted Pyramid: Upper Interior: Hidden Room in the Inverted Pyramid":
        "Sand: Hidden Room in the Inverted Pyramid",
    # Sheet rendered the "{Jaxi}" name template as an empty placeholder.
    "Sand Kingdom: Welcome Back, !":
        "Sand: Welcome Back, Jaxi!",
}

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _s(v) -> str:
    """Cell value → stripped string ('' for None)."""
    return "" if v is None else str(v).strip()


def _split_cell(cell: str) -> list[str]:
    """Split a comma-separated cell, strip whitespace, drop empty strings."""
    return [v.strip() for v in cell.split(",") if v.strip()]


def _norm_capture(name: str) -> str:
    return CAPTURE_NAME_MAP.get(name, name)


def _norm_jump(raw: str) -> str | None:
    v = raw.strip()
    if not v:
        return None
    if v not in JUMP_HEIGHT_MAP:
        raise ValueError(f"Unknown jump height: {v!r}")
    return JUMP_HEIGHT_MAP[v]


def _norm_cap_throws(raw: str) -> list[str]:
    result: list[str] = []
    for tok in _split_cell(raw):
        if tok not in CAP_THROW_MAP:
            raise ValueError(f"Unknown cap throw: {tok!r}")
        result.append(CAP_THROW_MAP[tok])
    return result


def _norm_other_required(raw: str) -> list[str]:
    result: list[str] = []
    for tok in _split_cell(raw):
        if tok not in OTHER_REQUIRED_MAP:
            raise ValueError(f"Unknown other-required term: {tok!r}")
        mapped = OTHER_REQUIRED_MAP[tok]
        if mapped is not None:          # None == "None" sentinel → skip
            result.append(mapped)
    return result


def _parse_capture_cells(cells: list[str]) -> tuple[list[list[str]], bool, bool]:
    """
    Parse the three col-C sub-cells of a moon block.

    Returns (capture_groups, capture_optional, locked_default_capture):
      capture_groups          OR of AND-groups (normalised AP names)
      capture_optional        True if a literal "None" sub-cell gives a no-capture path
      locked_default_capture  True if any sub-cell is the LOCKED_SENTINEL
    """
    groups: list[list[str]] = []
    optional = False
    locked = False
    for cell in cells:
        v = cell.strip()
        if not v:
            continue
        if v == LOCKED_SENTINEL:
            locked = True
            continue
        if v.lower() == "none":
            optional = True
            continue
        group = [_norm_capture(tok) for tok in _split_cell(v)]
        if group and group not in groups:
            groups.append(group)
    return groups, optional, locked


# ─────────────────────────────────────────────────────────────────────────────
# Workbook → rows (0-indexed list-of-lists, index = column-1; A..J)
# ─────────────────────────────────────────────────────────────────────────────

def _load_rows(xlsx_path: Path) -> list[list]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET_NAME!r} not found in {xlsx_path}")
    ws = wb[SHEET_NAME]
    # cols A..J → indices 0..9.  Merged cells: value only in the top-left cell,
    # the rest read as None — exactly what the capture AND/OR logic relies on.
    return [[ws.cell(r, c).value for c in range(1, 11)] for r in range(1, ws.max_row + 1)]


def _moon_starts(rows: list[list]) -> list[int]:
    """Block start = a row whose name (col B) is set and label (col D) == jump-height."""
    return [
        i for i in range(2, len(rows))
        if _s(rows[i][1]) and _s(rows[i][3]) == "Minimum Jump Height:"
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Locations.json reverse lookup
# ─────────────────────────────────────────────────────────────────────────────

def _build_location_lookup(data_dir: Path) -> dict[str, str]:
    """short_name → canonical_location_name for all non-Capture locations."""
    locs: list[dict] = json.loads((data_dir / "locations.json").read_text(encoding="utf-8"))
    lookup: dict[str, str] = {}
    for loc in locs:
        name: str = loc["name"]
        if name.startswith("Capture:"):
            continue
        if ": " in name:
            lookup[name.split(": ", 1)[1]] = name
    return lookup


def _load_capture_items(data_dir: Path) -> set[str]:
    """All AP capture item names (for validating sheet capture tokens)."""
    data = json.loads((data_dir / "items.json").read_text(encoding="utf-8"))
    items = data["items"] if isinstance(data, dict) and "items" in data else data
    caps: set[str] = set()
    for it in items:
        cat = it.get("category")
        cats = cat if isinstance(cat, list) else [cat]
        if "Capture" in cats:
            caps.add(it["name"])
    return caps


def _reconcile(name: str, short_lookup: dict[str, str]) -> str | None:
    if name in CSV_NAME_OVERRIDES:
        return CSV_NAME_OVERRIDES[name]
    parts = name.split(": ", 1)
    if len(parts) != 2:
        return None
    return short_lookup.get(parts[1].strip())


# ─────────────────────────────────────────────────────────────────────────────
# Main parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_rows(rows: list[list]) -> tuple[list[dict], list[str]]:
    records: list[dict] = []
    errors: list[str] = []

    for idx in _moon_starts(rows):
        if idx + 2 >= len(rows):
            errors.append(f"Row {idx}: truncated block for {rows[idx][1]!r}")
            continue

        row_jump, row_cap, row_other = rows[idx], rows[idx + 1], rows[idx + 2]
        name = _s(row_jump[1])

        capture_groups, capture_optional, locked = _parse_capture_cells(
            [_s(row_jump[2]), _s(row_cap[2]), _s(row_other[2])]
        )

        methods: dict[str, dict | None] = {}
        for m_idx, m_key in enumerate(("1", "2", "3", "4", "5")):
            col = 4 + m_idx  # cols E-I → indices 4-8
            jraw = _s(row_jump[col]) if col < len(row_jump) else ""
            craw = _s(row_cap[col]) if col < len(row_cap) else ""
            oraw = _s(row_other[col]) if col < len(row_other) else ""

            if not jraw and not craw and not oraw:
                methods[m_key] = None
                continue

            try:
                method = {
                    "jump_height":    _norm_jump(jraw),
                    "cap_throws":     _norm_cap_throws(craw),
                    "other_required": _norm_other_required(oraw),
                }
            except ValueError as e:
                errors.append(f"{name} method {m_key}: {e}")
                method = {"jump_height": None, "cap_throws": [], "other_required": []}
            methods[m_key] = method

        records.append({
            "csv_name":               name,
            "capture_groups":         capture_groups,
            "capture_optional":       capture_optional,
            "locked_default_capture": locked,
            "methods":                methods,
        })

    return records, errors


# ─────────────────────────────────────────────────────────────────────────────
# Subarea builder
# ─────────────────────────────────────────────────────────────────────────────

# Short kingdom name (the location-name prefix, e.g. "Snow") → full CSV kingdom
# name (e.g. "Snow Kingdom"). Reverse of KINGDOM_PREFIX_MAP.
SHORT_TO_FULL_KINGDOM: dict[str, str] = {
    short: full for full, short in KINGDOM_PREFIX_MAP.items()
}


def _row_kingdom(loc_name: str | None, sheet_kingdom: str | None) -> str | None:
    """Resolve a subarea moon's kingdom.

    The reconciled location name carries an authoritative short-kingdom prefix
    (e.g. "Snow: The Icicle Barrier" → Snow Kingdom). Prefer it over sheet order:
    some subarea blocks (Shiveria Town, Class A Race) appear in the sheet *before*
    their kingdom's first overworld moon, so "most-recently-seen main kingdom"
    misassigns them. Fall back to sheet order only for unreconciled moons.
    """
    if loc_name and ": " in loc_name:
        full = SHORT_TO_FULL_KINGDOM.get(loc_name.split(": ", 1)[0].strip())
        if full:
            return full
    return sheet_kingdom


def build_subareas(rows: list[list], records: list[dict],
                   short_lookup: dict[str, str],
                   entrance_stages: dict[str, dict] | None = None) -> dict[str, dict]:
    # entrance_stages is the authoritative stage topology (entrance_stages.json).
    # A prefix that spans >1 kingdom is split into per-kingdom "(ShortKingdom)"
    # entries ONLY when that topology actually has distinct per-kingdom stages
    # (Costume Room, Sphynx Treasure Vault). One-stage / two-entrance subareas
    # reached from two kingdoms via a warp painting (Picture Match (Goomba)) have
    # only the bare merged key in the topology and stay merged. When entrance_stages
    # is None the importer can't tell genuine splits from cross-kingdom entrances,
    # so it keeps everything merged (the entrance_logic round-trip filter then
    # drops any merged member that isn't resolvable — fail-safe, never one-way).
    entrance_stages = entrance_stages or {}
    entrance_keys = set(entrance_stages)

    # Sheet-order fallback: most-recently-seen main kingdom per subarea moon.
    current_kingdom: str | None = None
    sheet_kingdom: dict[str, str] = {}   # full csv_name -> kingdom (fallback only)
    for idx in _moon_starts(rows):
        csv_name = _s(rows[idx][1])
        parts = csv_name.split(": ", 1)
        if len(parts) != 2:
            continue
        prefix = parts[0].strip()
        if prefix in MAIN_KINGDOM_PREFIXES:
            current_kingdom = prefix
        elif current_kingdom:
            sheet_kingdom[csv_name] = current_kingdom

    # Pass 1: resolve each subarea moon's prefix + kingdom, and tally which
    # kingdoms each prefix spans. A prefix that spans >1 kingdom is "split" — its
    # entries get a "(ShortKingdom)" suffix so each kingdom's distinct subarea
    # stays a separate, entrance-shuffle-resolvable entry (Costume Room,
    # Sphynx Treasure Vault). See entrance_logic.is_round_trippable.
    resolved: list[tuple[str, str, str | None, str | None]] = []  # prefix, csv, kingdom, loc
    prefix_kingdoms: dict[str, set[str]] = {}
    for rec in records:
        csv_name = rec["csv_name"]
        parts = csv_name.split(": ", 1)
        if len(parts) != 2:
            continue
        prefix = parts[0].strip()
        if prefix in MAIN_KINGDOM_PREFIXES:
            continue
        loc = _reconcile(csv_name, short_lookup)
        kingdom = _row_kingdom(loc, sheet_kingdom.get(csv_name))
        resolved.append((prefix, csv_name, kingdom, loc))
        if kingdom:
            prefix_kingdoms.setdefault(prefix, set()).add(kingdom)

    split_prefixes = {p for p, ks in prefix_kingdoms.items() if len(ks) > 1}

    subareas: dict[str, dict] = {}
    for prefix, csv_name, kingdom, loc in resolved:
        key = prefix
        if prefix in split_prefixes and kingdom:
            cand = f"{prefix} ({KINGDOM_PREFIX_MAP.get(kingdom, kingdom)})"
            if cand in entrance_keys:
                key = cand
        # For subareas that have a door in the topology, the entrance_stages
        # kingdom is authoritative (it gates the door) — adopt it. Cross-kingdom
        # merged subareas (Picture Match (Goomba)) carry the topology's "owning"
        # kingdom there rather than whichever moon happened to resolve first.
        entry_kingdom = entrance_stages.get(key, {}).get("kingdom") or kingdom
        entry = subareas.setdefault(key, {
            "kingdom":        entry_kingdom,
            "csv_names":      [],
            "location_names": [],
        })
        entry["csv_names"].append(csv_name)
        if loc:
            entry["location_names"].append(loc)
    return subareas


# ─────────────────────────────────────────────────────────────────────────────
# Report
# ─────────────────────────────────────────────────────────────────────────────

def _print_report(records: list[dict], short_lookup: dict[str, str],
                  capture_items: set[str], parse_errors: list[str]) -> None:
    matched = unmatched = 0
    unmatched_names: list[str] = []
    or_moons: list[str] = []
    unknown_caps: dict[str, list[str]] = {}

    for rec in records:
        name = rec["csv_name"]
        if _reconcile(name, short_lookup):
            matched += 1
        else:
            unmatched += 1
            unmatched_names.append(name)
        if len(rec["capture_groups"]) > 1 or (rec["capture_groups"] and rec["capture_optional"]):
            or_moons.append(name)
        for grp in rec["capture_groups"]:
            for cap in grp:
                if cap not in capture_items:
                    unknown_caps.setdefault(cap, []).append(name)

    bar = "=" * 64
    print(f"\n{bar}")
    print(f"Moon blocks parsed             : {len(records)}")
    print(f"Matched to locations.json      : {matched}")
    print(f"Unmatched (not yet a location) : {unmatched}")
    print(f"Capture OR / optional moons     : {len(or_moons)}")

    if unknown_caps:
        print(f"\nCapture tokens NOT in items.json ({len(unknown_caps)}):")
        for cap, moons in sorted(unknown_caps.items()):
            print(f"  {cap!r}  (e.g. {moons[0]})")

    if or_moons:
        print(f"\nCapture OR / optional moons ({len(or_moons)}) — verify AND/OR split:")
        for n in or_moons:
            print(f"  {n}")

    if unmatched_names:
        print(f"\nUnmatched names ({unmatched}):")
        for n in unmatched_names:
            print(f"  {n}")

    if parse_errors:
        print(f"\nParse errors ({len(parse_errors)}):")
        for e in parse_errors:
            print(f"  {e}")
    print(bar)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.exit(f"Workbook not found: {xlsx_path}")

    print(f"Parsing: {xlsx_path}  (sheet {SHEET_NAME!r})")
    rows = _load_rows(xlsx_path)
    records, parse_errors = parse_rows(rows)
    print(f"  {len(records)} moon blocks, {len(parse_errors)} parse errors")

    short_lookup = _build_location_lookup(DATA_DIR)
    capture_items = _load_capture_items(DATA_DIR)
    print(f"  {len(short_lookup)} matchable locations, {len(capture_items)} AP captures")

    for rec in records:
        rec["location_name"] = _reconcile(rec["csv_name"], short_lookup)

    entrance_stages: dict[str, dict] | None = None
    if ENTRANCE_STAGES.exists():
        entrance_stages = json.loads(ENTRANCE_STAGES.read_text(encoding="utf-8"))
        print(f"  {len(entrance_stages)} entrance-stage keys (split topology)")
    else:
        print("  entrance_stages.json absent — subareas kept merged (no split)")
    subareas = build_subareas(rows, records, short_lookup, entrance_stages)

    # moon_requirements keyed by sheet name (drop the temp csv_name field)
    requirements: dict[str, dict] = {}
    for rec in records:
        key = rec.pop("csv_name")
        requirements[key] = rec

    OUT_REQUIREMENTS.write_text(
        json.dumps(requirements, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  -> {OUT_REQUIREMENTS}")
    OUT_SUBAREAS.write_text(
        json.dumps(subareas, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  -> {OUT_SUBAREAS}")

    # restore csv_name for the report helper
    report_records = [{"csv_name": k, **v} for k, v in requirements.items()]
    _print_report(report_records, short_lookup, capture_items, parse_errors)


if __name__ == "__main__":
    main()
